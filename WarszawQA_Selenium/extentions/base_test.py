#!/usr/bin/env python
#-*- coding: utf-8 -*-

# Example code for WarszawQA
# created by okulynyak@gmail.com

import datetime
import calendar
import codecs
import random
import csv
import string
import sys
import unicodedata
import time
import os
import xlrd
import ftplib
import io
import requests
from openpyxl import load_workbook

#line needed to send utf-8 characters to database
os.environ["NLS_LANG"] = ".UTF8"

class Base_test(object):
    """
    This class for methods that allows to get access by automated scripts
    to OS, FTP, EXCELL, DB, CSV and another functions.
    Methods are grouped in sub-classes by functionality
    Each method should be provided by short comment of it functionality
    """

    def generate_nonexistant_number_in_db(
            self,data,table,row,minimal_lenght,maximal_lenght):
        """
        This method returns unique (nonexistant in defined table and column)
        smallest number in range (defined by minimal_lenght and maximal_lenght)
        """
        result = 0
        start = pow(10,minimal_lenght)-(9*pow(10,minimal_lenght-1))
        end = pow(10,maximal_lenght)-1

        for i in range(start,end):
            if not self.DB_api().check_query(
                data["DB credentials"],
                "select * from {0} where {1}='{2}'".format(table,row,i)):
                result = i
                break
        return result


    class Excell_api():
        """
        All methods are placed here for access and do different actions
        with MS Excell 2007 file type.
        """
        def read_xls(self, path, environment, testnames):
            #this function return dictionary fron 'Environment variables' sheet and
            #sheet named as each ekement in list argument 'testnames'
            d = {}
            wb = xlrd.open_workbook(path+"\config_{0}.xlsx".format(environment))
            sh = wb.sheet_by_name("Environment variables")
            i=0
            while(True):
                try:
                    cell_value_class = str(sh.cell(0,i).value)
                    cell_value_id = str(sh.cell(1,i).value).encode('utf-8')
                    d[cell_value_class] = cell_value_id
                    i+=1
                except: break
            for testname in testnames:
                sh = wb.sheet_by_name(testname)
                i=0
                while(True):
                    try:
                        cell_value_class = str(sh.cell(0,i).value)
                        try:
                            cell_value_id = str((sh.cell(1,i).value).encode('utf-8'))
                        except:
                            cell_value_id = str(int(sh.cell(1,i).value))
                        d[cell_value_class] = cell_value_id
                        i+=1
                    except:
                        try:
                            cell_value_class = str(sh.cell(0,i).value)
                            result=sh.cell(1,i).value
                            cell_value_id = str((sh.cell(1,i).value).encode('utf-8'))
                            d[cell_value_class] = cell_value_id.decode('utf-8')
                            i+=1
                        except:
                                break
            return d

    class TXT_api():
        """
        All methods for work with TXT files here
        """
        def config_file_to_dictionary(self, file_path):
            """
            read txt file with lines like {key=value
            and return dictionaty according to information
            """
            opened_file = open(file_path,'r')
            dictionary_of_values={}
            for line in opened_file:
                if "=" in line:
                    key = line.split("=")[0].strip()
                    dictionary_of_values[key] = str(line.split("=")[1]).strip()
            return dictionary_of_values

    class DB_api():
        """
        All methods are placed here for access and do
        different actions with SQL Oracle DB.
        """
        
        def check_query(self, database_credentials, query):
            """
            this method checks if query returns one or more results
            if it does - returns True, if it does not - returns False
            """
            dictionary = self.make_list_of_dictionaries_from_db(database_credentials, query)
            if(len(dictionary) == 0):
                return False
            else:
                return True


    class FTP_api():
        """
        All methods are placed here for access and do different actions with FTP (SFTP).
        """

    class OS_api():
        """
        All methods are placed here for access and do different actions Microsoft Windows OS.
        """

    class CSV_api():
        """
        All methods are placed here for access and do different actions
        with CSV file types.
        """
        def read_csv_to_dict_table(self, path_to_file, separator=";"):
            '''Read csv file and return dictionary for each line in referens of first line
            '''
            #path_to_file=path+filename
            file_n = ''
            try:
                file_n = codecs.open(path_to_file,
                                     'r',
                                     encoding='utf-8-sig',
                                     errors='ignore')
                #lines = [line.strip() for line in file_n]
            except IOError:
                print ("ERROR: Nie można odnaleźć pliku - " + path_to_file)
                current_time = datetime.datetime.now()
                print ("STATUS TESTU: FAILED")
                print ("STOP TESTU:  %s" % current_time.isoformat())
                print ("=======================================================================")
                exit(0)
            finally:
                rows = file_n.readlines() # read everything in the file
                rows_amount = len(rows)
            rows_as_tables=[]
            for row in rows: #separate rows into tables
                row_as_table=row.split(separator)
                rows_as_tables.append(row_as_table)
            temp_dict={}
            dict_table=[]
            iterator=0
            for row in rows_as_tables: 
                if(iterator!=0):
                    table_iterator=0
                    for field in row:
                        if(field!='\n'):
                            temp_dict[(rows_as_tables[0][table_iterator])]=str(field)
                            table_iterator+=1
                    dict_table.append(temp_dict)
                    temp_dict={}
                iterator+=1
            file_n.close()
            return dict_table
